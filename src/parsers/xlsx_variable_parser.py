import os
import re
import logging
from typing import List, Dict, Optional, Any, Tuple
from dataclasses import dataclass

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


@dataclass
class PLCVariable:
    name: str
    logical_address: str
    data_type: str
    db_number: Optional[int] = None
    byte_offset: Optional[int] = None
    bit_offset: Optional[int] = None
    area: Optional[str] = None
    description: str = ""
    access_type: str = "read"
    
    def to_dict(self) -> Dict[str, Any]:
        return {
            'name': self.name,
            'logical_address': self.logical_address,
            'data_type': self.data_type,
            'db_number': self.db_number,
            'byte_offset': self.byte_offset,
            'bit_offset': self.bit_offset,
            'area': self.area,
            'description': self.description,
            'access_type': self.access_type
        }


class XLSXVariableParser:
    AREA_MAP = {
        'M': 'M',
        'I': 'I',
        'Q': 'Q',
        'E': 'I',
        'A': 'Q',
        'DB': 'DB',
        'INPUT': 'I',
        'OUTPUT': 'Q',
        'MEMORY': 'M',
    }

    TYPE_MAP = {
        'BOOL': 'BOOL',
        'BIT': 'BOOL',
        'BYTE': 'BYTE',
        'WORD': 'WORD',
        'DWORD': 'DWORD',
        'INT': 'INT',
        'DINT': 'DINT',
        'REAL': 'REAL',
        'LREAL': 'LREAL',
        'STRING': 'STRING',
        'CHAR': 'CHAR',
        'SINT': 'INT',
        'USINT': 'INT',
        'UINT': 'INT',
        'UDINT': 'DINT',
        'TIMER': 'TIMER',
        'COUNTER': 'COUNTER',
    }

    NAME_KEYS = ['名称', 'Name', '变量名', 'Tag Name', '标签名', 'Tag', 'Variable']
    ADDRESS_KEYS = ['地址', 'Address', '逻辑地址', 'Logical Address', '偏移地址', 'Offset', 'Addr']
    TYPE_KEYS = ['类型', 'Type', '数据类型', 'Data Type', 'DataType']
    DESCRIPTION_KEYS = ['描述', 'Description', '说明', 'Comment', '备注', '注释']
    ACCESS_KEYS = ['访问类型', 'Access Type', 'Access', '读写', 'RW', 'ReadOnly', 'WriteOnly']

    def __init__(self):
        self.variables: Dict[str, PLCVariable] = {}
        self._address_patterns = self._build_address_patterns()

    def _build_address_patterns(self) -> List[Tuple[re.Pattern, str]]:
        patterns = [
            (re.compile(r'%DB(\d+)\.DBX(\d+)\.(\d+)'), 'DB_BIT'),
            (re.compile(r'%DB(\d+)\.DBB(\d+)'), 'DB_BYTE'),
            (re.compile(r'%DB(\d+)\.DBW(\d+)'), 'DB_WORD'),
            (re.compile(r'%DB(\d+)\.DBD(\d+)'), 'DB_DWORD'),
            (re.compile(r'%DB(\d+)\.(\d+)\.(\d+)'), 'DB_BIT_ALT'),
            (re.compile(r'%DB(\d+)\.(\d+)'), 'DB_WORD_ALT'),
            (re.compile(r'DB(\d+)\.DBX(\d+)\.(\d+)'), 'DB_BIT_NO_PERCENT'),
            (re.compile(r'DB(\d+)\.DBB(\d+)'), 'DB_BYTE_NO_PERCENT'),
            (re.compile(r'DB(\d+)\.DBW(\d+)'), 'DB_WORD_NO_PERCENT'),
            (re.compile(r'DB(\d+)\.DBD(\d+)'), 'DB_DWORD_NO_PERCENT'),
            (re.compile(r'DB(\d+)\.(\d+)\.(\d+)'), 'DB_BIT_SIMPLE'),
            (re.compile(r'DB(\d+)\.(\d+)'), 'DB_WORD_SIMPLE'),
            (re.compile(r'(\d+)\.(\d+)\.(\d+)'), 'DB_BIT_NUMERIC'),
            (re.compile(r'(\d+)\.(\d+)'), 'DB_WORD_NUMERIC'),
            (re.compile(r'DB(\d+)'), 'DB_ONLY'),
            (re.compile(r'%I(\d+)\.(\d+)'), 'INPUT_BIT'),
            (re.compile(r'%IB(\d+)'), 'INPUT_BYTE'),
            (re.compile(r'%IW(\d+)'), 'INPUT_WORD'),
            (re.compile(r'%ID(\d+)'), 'INPUT_DWORD'),
            (re.compile(r'%I(\d+)'), 'INPUT_BYTE'),
            (re.compile(r'I(\d+)\.(\d+)'), 'INPUT_BIT_NO_PERCENT'),
            (re.compile(r'IB(\d+)'), 'INPUT_BYTE_NO_PERCENT'),
            (re.compile(r'I(\d+)'), 'INPUT_BYTE_NO_PERCENT'),
            (re.compile(r'%Q(\d+)\.(\d+)'), 'OUTPUT_BIT'),
            (re.compile(r'%QB(\d+)'), 'OUTPUT_BYTE'),
            (re.compile(r'%QW(\d+)'), 'OUTPUT_WORD'),
            (re.compile(r'%QD(\d+)'), 'OUTPUT_DWORD'),
            (re.compile(r'%Q(\d+)'), 'OUTPUT_BYTE'),
            (re.compile(r'Q(\d+)\.(\d+)'), 'OUTPUT_BIT_NO_PERCENT'),
            (re.compile(r'QB(\d+)'), 'OUTPUT_BYTE_NO_PERCENT'),
            (re.compile(r'Q(\d+)'), 'OUTPUT_BYTE_NO_PERCENT'),
            (re.compile(r'%M(\d+)\.(\d+)'), 'MEMORY_BIT'),
            (re.compile(r'%MB(\d+)'), 'MEMORY_BYTE'),
            (re.compile(r'%MW(\d+)'), 'MEMORY_WORD'),
            (re.compile(r'%MD(\d+)'), 'MEMORY_DWORD'),
            (re.compile(r'%M(\d+)'), 'MEMORY_BYTE'),
            (re.compile(r'M(\d+)\.(\d+)'), 'MEMORY_BIT_NO_PERCENT'),
            (re.compile(r'MB(\d+)'), 'MEMORY_BYTE_NO_PERCENT'),
            (re.compile(r'M(\d+)'), 'MEMORY_BYTE_NO_PERCENT'),
            (re.compile(r'%T(\d+)'), 'TIMER'),
            (re.compile(r'T(\d+)'), 'TIMER'),
            (re.compile(r'%C(\d+)'), 'COUNTER'),
            (re.compile(r'C(\d+)'), 'COUNTER'),
        ]
        return patterns

    def parse_xlsx(self, file_path: str, sheet_name: str = None) -> Dict[str, PLCVariable]:
        if not os.path.exists(file_path):
            logger.error(f"XLSX file not found: {file_path}")
            raise FileNotFoundError(f"XLSX file not found: {file_path}")

        try:
            return self._parse_with_openpyxl(file_path, sheet_name)
        except ImportError:
            logger.warning("openpyxl not installed, trying xlrd...")
            return self._parse_with_xlrd(file_path, sheet_name)

    def _parse_with_openpyxl(self, file_path: str, sheet_name: str = None) -> Dict[str, PLCVariable]:
        import openpyxl
        
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        ws = self._get_worksheet(wb, sheet_name)
        
        if ws is None:
            logger.error(f"Worksheet not found: {sheet_name}")
            wb.close()
            return {}

        headers = []
        header_indices = {}
        self.variables = {}

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_idx == 1:
                headers, header_indices = self._parse_headers(row)
                logger.info(f"Found headers: {headers}")
                continue

            if not row or all(cell is None for cell in row):
                continue

            try:
                var = self._parse_row_with_indices(row, header_indices)
                if var:
                    self.variables[var.name] = var
            except Exception as e:
                logger.warning(f"Error parsing row {row_idx}: {e}")

        wb.close()
        logger.info(f"Successfully parsed {len(self.variables)} variables from {file_path}")
        return self.variables

    def _parse_with_xlrd(self, file_path: str, sheet_name: str = None) -> Dict[str, PLCVariable]:
        try:
            import xlrd
        except ImportError:
            logger.error("xlrd not installed either, cannot parse xlsx file")
            return {}

        wb = xlrd.open_workbook(file_path)
        ws = self._get_worksheet_xlrd(wb, sheet_name)
        
        if ws is None:
            logger.error(f"Worksheet not found: {sheet_name}")
            return {}

        headers = []
        header_indices = {}
        self.variables = {}

        for row_idx in range(ws.nrows):
            row = ws.row_values(row_idx)
            if row_idx == 0:
                headers, header_indices = self._parse_headers(row)
                logger.info(f"Found headers: {headers}")
                continue

            if not row or all(cell == '' for cell in row):
                continue

            try:
                var = self._parse_row_with_indices(row, header_indices)
                if var:
                    self.variables[var.name] = var
            except Exception as e:
                logger.warning(f"Error parsing row {row_idx + 1}: {e}")

        logger.info(f"Successfully parsed {len(self.variables)} variables from {file_path}")
        return self.variables

    def _get_worksheet(self, wb, sheet_name: Optional[str]):
        if sheet_name:
            if sheet_name in wb.sheetnames:
                return wb[sheet_name]
            logger.warning(f"Sheet '{sheet_name}' not found, using active sheet")
            return wb.active
        return wb.active

    def _get_worksheet_xlrd(self, wb, sheet_name: Optional[str]):
        if sheet_name:
            try:
                return wb.sheet_by_name(sheet_name)
            except Exception:
                logger.warning(f"Sheet '{sheet_name}' not found, using first sheet")
                return wb.sheet_by_index(0)
        return wb.sheet_by_index(0)

    def _parse_headers(self, row: List[Any]) -> Tuple[List[str], Dict[str, int]]:
        headers = []
        indices = {}
        
        for idx, cell in enumerate(row):
            header = str(cell).strip() if cell else f"col_{idx}"
            headers.append(header)
            
            header_lower = header.lower()
            
            # Only set index if not already set to prevent overwriting
            if 'name' not in indices and any(key.lower() in header_lower for key in self.NAME_KEYS):
                indices['name'] = idx
            elif 'address' not in indices and any(key.lower() in header_lower for key in self.ADDRESS_KEYS):
                indices['address'] = idx
            elif 'type' not in indices and any(key.lower() in header_lower for key in self.TYPE_KEYS):
                indices['type'] = idx
            elif 'description' not in indices and any(key.lower() in header_lower for key in self.DESCRIPTION_KEYS):
                indices['description'] = idx
            elif 'access' not in indices and any(key.lower() in header_lower for key in self.ACCESS_KEYS):
                indices['access'] = idx

        return headers, indices

    def _parse_row_with_indices(self, row: List[Any], indices: Dict[str, int]) -> Optional[PLCVariable]:
        name = self._get_cell_value(row, indices, 'name')
        if not name:
            return None

        name = str(name).strip()

        address = self._get_cell_value(row, indices, 'address')
        data_type = self._get_cell_value(row, indices, 'type', 'BOOL')
        description = self._get_cell_value(row, indices, 'description', '')
        access_type = self._get_cell_value(row, indices, 'access', 'read')

        if address:
            parsed = self._parse_address(str(address))
        else:
            parsed = {'area': None, 'db': None, 'offset': 0, 'bit': 0}

        return PLCVariable(
            name=name,
            logical_address=str(address) if address else '',
            data_type=self._normalize_type(str(data_type)),
            db_number=parsed.get('db'),
            byte_offset=parsed.get('offset', 0),
            bit_offset=parsed.get('bit', 0),
            area=parsed.get('area'),
            description=str(description) if description else '',
            access_type=str(access_type).lower() if access_type else 'read'
        )

    def _get_cell_value(self, row: List[Any], indices: Dict[str, int], key: str, default: Any = None) -> Any:
        idx = indices.get(key)
        if idx is None or idx >= len(row):
            return default
        
        value = row[idx]
        if value is None:
            return default
        
        str_value = str(value).strip()
        return str_value if str_value else default

    def _parse_address(self, address: str) -> Dict[str, Any]:
        address = address.strip().upper()

        for pattern, pattern_type in self._address_patterns:
            match = pattern.match(address)
            if match:
                return self._parse_pattern_match(match, pattern_type)

        return {'area': None, 'db': None, 'offset': 0, 'bit': 0}

    def _parse_pattern_match(self, match: re.Match, pattern_type: str) -> Dict[str, Any]:
        groups = match.groups()
        
        if pattern_type.startswith('DB'):
            result = {'area': 'DB', 'db': int(groups[0])}
            if pattern_type in ('DB_BIT', 'DB_BIT_ALT', 'DB_BIT_NO_PERCENT', 'DB_BIT_SIMPLE'):
                result['offset'] = int(groups[1])
                result['bit'] = int(groups[2])
            elif pattern_type in ('DB_BYTE', 'DB_BYTE_NO_PERCENT'):
                result['offset'] = int(groups[1])
                result['bit'] = 0
            elif pattern_type in ('DB_WORD', 'DB_WORD_ALT', 'DB_WORD_NO_PERCENT', 'DB_WORD_SIMPLE'):
                result['offset'] = int(groups[1])
                result['bit'] = 0
            elif pattern_type in ('DB_DWORD', 'DB_DWORD_NO_PERCENT'):
                result['offset'] = int(groups[1])
                result['bit'] = 0
            elif pattern_type == 'DB_BIT_NUMERIC':
                result['db'] = int(groups[0])
                result['offset'] = int(groups[1])
                result['bit'] = int(groups[2])
            elif pattern_type == 'DB_WORD_NUMERIC':
                result['db'] = int(groups[0])
                result['offset'] = int(groups[1])
                result['bit'] = 0
            elif pattern_type == 'DB_ONLY':
                result['offset'] = 0
                result['bit'] = 0
            return result
        
        elif pattern_type.startswith('INPUT'):
            result = {'area': 'I', 'db': None}
            if pattern_type.endswith('BIT') or pattern_type.endswith('BIT_NO_PERCENT'):
                result['offset'] = int(groups[0])
                result['bit'] = int(groups[1]) if len(groups) > 1 else 0
            else:
                result['offset'] = int(groups[0])
                result['bit'] = 0
            return result
        
        elif pattern_type.startswith('OUTPUT'):
            result = {'area': 'Q', 'db': None}
            if pattern_type.endswith('BIT') or pattern_type.endswith('BIT_NO_PERCENT'):
                result['offset'] = int(groups[0])
                result['bit'] = int(groups[1]) if len(groups) > 1 else 0
            else:
                result['offset'] = int(groups[0])
                result['bit'] = 0
            return result
        
        elif pattern_type.startswith('MEMORY'):
            result = {'area': 'M', 'db': None}
            if pattern_type.endswith('BIT') or pattern_type.endswith('BIT_NO_PERCENT'):
                result['offset'] = int(groups[0])
                result['bit'] = int(groups[1]) if len(groups) > 1 else 0
            else:
                result['offset'] = int(groups[0])
                result['bit'] = 0
            return result
        
        elif pattern_type == 'TIMER':
            return {'area': 'T', 'db': None, 'offset': int(groups[0]), 'bit': 0}
        
        elif pattern_type == 'COUNTER':
            return {'area': 'C', 'db': None, 'offset': int(groups[0]), 'bit': 0}
        
        return {'area': None, 'db': None, 'offset': 0, 'bit': 0}

    def _normalize_type(self, data_type: str) -> str:
        data_type = data_type.upper().strip()

        for key, normalized in self.TYPE_MAP.items():
            if key in data_type or data_type in key:
                return normalized

        return 'BOOL'

    def get_variables_by_area(self, area: str) -> List[PLCVariable]:
        area = area.upper()
        mapped_area = self.AREA_MAP.get(area, area)
        return [v for v in self.variables.values() if v.area == mapped_area]

    def get_variables_by_db(self, db_number: int) -> List[PLCVariable]:
        return [v for v in self.variables.values() if v.db_number == db_number]

    def get_m_variables(self) -> List[PLCVariable]:
        return self.get_variables_by_area('M')

    def get_i_variables(self) -> List[PLCVariable]:
        return self.get_variables_by_area('I')

    def get_q_variables(self) -> List[PLCVariable]:
        return self.get_variables_by_area('Q')

    def get_db_variables(self) -> List[PLCVariable]:
        return [v for v in self.variables.values() if v.db_number is not None]

    def export_to_config_format(self) -> Dict[str, Any]:
        config = {
            'db_variables': {},
            'm_variables': [],
            'i_variables': [],
            'q_variables': [],
            'total': len(self.variables)
        }

        for var in self.variables.values():
            var_dict = {
                'name': var.name,
                'offset': var.byte_offset,
                'bit': var.bit_offset,
                'type': var.data_type,
                'description': var.description
            }

            if var.db_number is not None:
                if var.db_number not in config['db_variables']:
                    config['db_variables'][var.db_number] = []
                config['db_variables'][var.db_number].append(var_dict)
            elif var.area == 'M':
                config['m_variables'].append(var_dict)
            elif var.area == 'I':
                config['i_variables'].append(var_dict)
            elif var.area == 'Q':
                config['q_variables'].append(var_dict)

        return config

    def get_summary(self) -> Dict[str, Any]:
        summary = {
            'total': len(self.variables),
            'by_area': {'I': 0, 'Q': 0, 'M': 0, 'DB': 0},
            'by_type': {},
            'db_count': {}
        }

        for var in self.variables.values():
            if var.area in summary['by_area']:
                summary['by_area'][var.area] += 1
            
            summary['by_type'][var.data_type] = summary['by_type'].get(var.data_type, 0) + 1
            
            if var.db_number is not None:
                summary['db_count'][var.db_number] = summary['db_count'].get(var.db_number, 0) + 1

        return summary

    def validate_variables(self) -> List[Dict[str, Any]]:
        errors = []
        
        for name, var in self.variables.items():
            if not var.logical_address:
                errors.append({
                    'type': 'missing_address',
                    'variable': name,
                    'message': '变量缺少地址'
                })
            
            if var.db_number is not None and var.byte_offset is None:
                errors.append({
                    'type': 'invalid_db_offset',
                    'variable': name,
                    'message': 'DB变量缺少偏移地址'
                })
        
        return errors


def parse_xlsx_variable_file(file_path: str, sheet_name: str = None) -> Dict[str, PLCVariable]:
    """解析XLSX格式的PLC变量文件"""
    parser = XLSXVariableParser()
    return parser.parse_xlsx(file_path, sheet_name)


def get_xlsx_variable_summary(file_path: str, sheet_name: str = None) -> Dict[str, Any]:
    """获取XLSX变量文件的摘要信息"""
    parser = XLSXVariableParser()
    parser.parse_xlsx(file_path, sheet_name)
    return parser.get_summary()


def validate_xlsx_variables(file_path: str, sheet_name: str = None) -> List[Dict[str, Any]]:
    """验证XLSX变量文件的有效性"""
    parser = XLSXVariableParser()
    parser.parse_xlsx(file_path, sheet_name)
    return parser.validate_variables()